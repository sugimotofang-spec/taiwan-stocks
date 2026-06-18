#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import sys, io, glob, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
import openpyxl

BASE = os.path.dirname(os.path.abspath(__file__))
path = os.path.join(BASE, "2026config0618.xlsx")

wb = openpyxl.load_workbook(path)
ws = wb.active

# 三支股票的正確更正
CORRECTIONS = {
    "3455": {  # 由田
        "role":  "⚔️ 七殺爆量遊擊隊【地劫⚠️】",
        "star":  "七殺（事業宮申・旺）⚠️地劫",
        "wuxing":"金/火（高波動設備）",
    },
    "3587": {  # 閎康
        "role":  "🌀 天機智慧策略股",
        "star":  "天機（遷移宮戌・利）",
        "wuxing":"木（智識策略型）",
    },
    "3580": {  # 友威科（維持太陰，確認正確）
        "role":  "🌙 太陰穩健財庫",
        "star":  "太陰（命宮辰・化權）★核心",
        "wuxing":"水（防守穩健型）",
    },
}

# 欄位對應（從之前掃描得知）
# A=1:代號  B=2:名稱  C=3:角色定位  D=4:命理星宿  E=5:五行屬性
COL_CODE   = 1
COL_ROLE   = 3
COL_STAR   = 4
COL_WUXING = 5

updated = 0
for r in range(1, ws.max_row + 1):
    code = str(ws.cell(r, COL_CODE).value or "").strip()
    if code in CORRECTIONS:
        c = CORRECTIONS[code]
        name = ws.cell(r, 2).value
        old_role  = ws.cell(r, COL_ROLE).value
        old_star  = ws.cell(r, COL_STAR).value
        old_wx    = ws.cell(r, COL_WUXING).value
        ws.cell(r, COL_ROLE).value   = c["role"]
        ws.cell(r, COL_STAR).value   = c["star"]
        ws.cell(r, COL_WUXING).value = c["wuxing"]
        print(f"row{r} {code}{name}:")
        print(f"  角色: {old_role} → {c['role']}")
        print(f"  星宿: {old_star} → {c['star']}")
        print(f"  五行: {old_wx} → {c['wuxing']}")
        updated += 1

out = path.replace("0618.xlsx", "0618b.xlsx")
wb.save(out)
print(f"\n更新 {updated} 行，儲存至: {out}")
